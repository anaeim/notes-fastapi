"""
xlsx_to_string — Convert an xlsx worksheet to an LLM-friendly markdown string.

Overview
--------
Reads an .xlsx file locally using openpyxl (no external API calls) and converts
one or more sheets into a compact, token-efficient string suitable for inclusion
in LLM prompts.

Output format
-------------
Each sheet is rendered as:

    === Sheet: <name> ===
    <preamble lines>          # optional — metadata rows above the column headers

    | Col1 | Col2 | Col3 |   # column header row
    | ---  | ---  | ---  |
    | val  | val  | val  |   # data rows

Preamble rows (sparse rows before the column-header row, e.g. a company name or
pay-period line) are rendered as pipe-separated text lines rather than table rows,
because they don't share the same column structure as the data.

Header detection
----------------
The column-header row is identified as the first row where every column is
populated (non-None cell count == total column count). If no fully-populated row
exists, the row with the most non-None values is used as a fallback. This reliably
separates sparse metadata preambles from the actual column headers.

Supported table variations
--------------------------
- Simple table: single header row followed by data rows, no merged cells.
- Complex table without merge: metadata preamble rows above the column headers,
  no merged cells in the data area.
- Complex table with merge: same as above, but with merged cells in the preamble
  rows. Merged cells are handled transparently — openpyxl returns None for all
  non-top-left cells of a merge range, so only the actual value is captured.

Cell formatting
---------------
- datetime / date  →  YYYY-MM-DD  (ISO 8601, compact)
- Multi-line text  →  newlines replaced with " / " (kept on one line)
- None / blank     →  empty string (no "null" or "None" text in output)
- All other types  →  str(), stripped of leading/trailing whitespace

Dependencies
------------
- openpyxl  (install: pip install openpyxl)

Public API
----------
    xlsx_to_string(file_path, sheet_name=None) -> str
        Convert an xlsx file to a string. Pass sheet_name to target a single
        sheet; omit it to include all sheets.

    sheet_to_string(ws) -> str
        Convert a single openpyxl Worksheet object to a string.
"""

import openpyxl
from datetime import datetime, date, time


def _fmt_time(t: time) -> str:
    """Format a time object, omitting seconds/microseconds when they are zero."""
    if t.second == 0 and t.microsecond == 0:
        return t.strftime("%H:%M")
    if t.microsecond == 0:
        return t.strftime("%H:%M:%S")
    return t.strftime("%H:%M:%S.%f").rstrip("0")


def _fmt(value) -> str:
    """Format a cell value to a compact string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Include time only when it carries real information (not midnight)
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.strftime("%Y-%m-%d")
        return f"{value.strftime('%Y-%m-%d')} {_fmt_time(value.time())}"
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return _fmt_time(value)
    # Flatten multi-line text; strip whitespace
    return str(value).strip().replace("\n", " / ").replace("\r", "")


def _is_empty_row(row: tuple) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _find_header_row_idx(rows: list) -> int:
    """
    Return the 0-based index of the column-header row.

    Strategy: the header row is the first row where every column is populated
    (non-None count == total columns). If no such row exists, fall back to the
    row with the most non-None values.
    """
    col_count = len(rows[0]) if rows else 0
    best_idx, best_count = 0, 0

    for i, row in enumerate(rows):
        count = sum(1 for v in row if v is not None)
        if count == col_count:
            return i
        if count > best_count:
            best_count, best_idx = count, i

    return best_idx


def sheet_to_string(ws) -> str:
    """Convert a single worksheet to an LLM-friendly markdown string."""
    rows = list(ws.iter_rows(values_only=True))

    # Drop trailing blank rows
    while rows and _is_empty_row(rows[-1]):
        rows.pop()

    if not rows:
        return "(empty)"

    header_idx = _find_header_row_idx(rows)
    parts = []

    # --- Preamble (rows before the column-header row) ---
    preamble = [rows[i] for i in range(header_idx) if not _is_empty_row(rows[i])]
    if preamble:
        for row in preamble:
            # Merged cells appear as None for non-top-left cells; skip them.
            values = [_fmt(v) for v in row if v is not None and str(v).strip()]
            if values:
                parts.append(" | ".join(values))
        parts.append("")  # blank line before table

    # --- Column headers ---
    headers = [_fmt(v) for v in rows[header_idx]]
    parts.append("| " + " | ".join(headers) + " |")
    parts.append("| " + " | ".join(["---"] * len(headers)) + " |")

    # --- Data rows ---
    for row in rows[header_idx + 1:]:
        if _is_empty_row(row):
            continue
        parts.append("| " + " | ".join(_fmt(v) for v in row) + " |")

    return "\n".join(parts)


def xlsx_to_string(file_path: str, sheet_name: str = None) -> str:
    """
    Convert an xlsx file to an LLM-friendly string.

    Args:
        file_path:  Path to the .xlsx file.
        sheet_name: Sheet to convert. If None, all sheets are included.

    Returns:
        A string with one section per sheet, each rendered as a markdown table
        with an optional preamble for metadata rows.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    names = [sheet_name] if sheet_name else wb.sheetnames

    sections = []
    for name in names:
        content = sheet_to_string(wb[name])
        sections.append(f"=== Sheet: {name} ===\n{content}")

    return "\n\n".join(sections)


# ---------------------------------------------------------------------------
# Quick smoke test
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import os
    # xlsx = "xlsx2_complex_with_merge.xlsx"
    # xlsx = "xlsx_complex_without_merge.xlsx"
    xlsx = "xlsx_simple.xlsx"


    path = os.path.join(os.path.dirname(__file__), "samples", xlsx)
    print(xlsx_to_string(path))
